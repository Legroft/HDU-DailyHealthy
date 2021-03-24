from selenium.webdriver import Chrome
from selenium.webdriver.chrome.options import Options
# 导入时间模块
import time
from openpyxl import load_workbook
import requests
from selenium.common.exceptions import NoSuchElementException, ElementClickInterceptedException,ElementNotInteractableException

# 获取驱动路径
# linux
DRIVER_PATH = './chromedriver'
# 浏览器设置
options = Options()
options.add_argument('--no-sandbox')
# 无头参数
options.add_argument('--headless')
options.add_argument('--disable-gpu')

def qmsg(qq, msg):
    if qq=='0':
        pass
    else:
        # qmsg消息推送
        url = "https://qmsg.zendee.cn/send/" + str(qq)
        res = requests.post(url=url, data={"msg": msg})


def daka(token, p, c, a, qq):
    js = 'window.localStorage.setItem("hduhelp_ncov_dailysign_token","' + token + '")'
    browser = Chrome(executable_path=DRIVER_PATH, options=options)
    # 访问url
    browser.get("https://healthcheckin.hduhelp.com/")
    # 窗口最大化
    browser.maximize_window()
    # 添加token
    browser.execute_script(js)
    # 刷新浏览器
    browser.refresh()
    time.sleep(3)
    print("正在执行"+token+"操作")
    # 点击 确认打卡 按钮
    print("正在点击确认抗疫答题按钮")
    try:
        browser.find_element_by_css_selector('.van-hairline--top.van-dialog__footer').click()
        print("点击确认抗疫答题按钮成功")
    except NoSuchElementException:
        pass
    time.sleep(3)
    print("正在点击确认打卡按钮")
    if browser.find_element_by_css_selector('.van-button.van-button--info.van-button--normal').is_enabled()==False:
        print(token+"今日已打卡")
        qmsg(qq, "今日已打卡！")
        browser.quit()
    else:
        browser.find_element_by_css_selector('.van-button.van-button--info.van-button--normal').click()
        print("点击确认打卡按钮成功")
        time.sleep(3)
        # 点击弹出的 确认 按钮
        try:
            print("正在点击确认按钮")
            browser.find_element_by_class_name('van-dialog__confirm').click()
            print("点击确认按钮成功")
        except (NoSuchElementException, ElementNotInteractableException):
            print(token+"授权地理位置时出错")
            qmsg(qq, "授权地理位置时出错")
            browser.quit()
            return
        time.sleep(3)
        # 点击 确认 手动填写位置按钮
        try:
            print("点击地理选择框")
            browser.find_element_by_class_name('van-field__control--right').click()
        except NoSuchElementException:
            qmsg(qq, "打卡失败，选择位置时出错X01")
            print(token+"打卡失败，选择位置时出错X01")
            browser.quit()
            return
        time.sleep(3)
        # 获取滑动选择框
        try:
            print("正在选择地理位置")
            pickers = browser.find_elements_by_class_name('van-picker-column__wrapper')
            for i in range(p):
                # 依次点击 直到选择了对应城市
                pickers[0].find_elements_by_class_name('van-picker-column__item')[i].click()
            for i in range(c):
                # 同上
                pickers[1].find_elements_by_class_name('van-picker-column__item')[i].click()
            for i in range(a):
                # 同上
                pickers[2].find_elements_by_class_name('van-picker-column__item')[i].click()
        except NoSuchElementException:
            qmsg(qq, "打卡失败，选择位置时出错！X02")
            print(token+"打卡失败，选择位置时出错X02")
            browser.quit()
            return
        # 点击 确认 地区选择按钮
        try:
            print("确认位置")
            browser.find_element_by_class_name('van-picker__confirm').click()
        except NoSuchElementException:
            qmsg(qq, "打卡失败，确认位置时出错！")
            print(token+"打卡失败，确认位置时出错")
            browser.quit()
            return 
        time.sleep(3)
        # 点击 确认打卡 按钮
        try:
            browser.find_element_by_css_selector('.van-button.van-button--info.van-button--normal').click()
        except NoSuchElementException:
            qmsg(qq, "打卡失败，确认打卡时出错！")
            print(token+"打卡失败，确认打卡时出错")
            browser.quit()
            return 
        time.sleep(3)
        # 点击 确认负责 按钮
        try:
            print("正在确认负责")
            browser.find_element_by_css_selector('.van-button.van-button--default.van-button--large.van-dialog__confirm.van'
                                             '-hairline--left').click()
            print(token+"打卡成功！")
            qmsg(qq,"打卡成功！")
        except NoSuchElementException:
            qmsg(qq, "打卡失败，确认负责时出错！")
            print(token+"打卡失败，确认负责时出错")
            browser.quit()
            return 
        # 退出窗口
        browser.quit()
    

wb = load_workbook('data.xlsx')
ws = wb.active
datas = []
for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
    datas.append(row)
for data in datas:
    [token, p, c, a, qq] = data
    daka(token, p, c, a, qq)
    time.sleep(3)
